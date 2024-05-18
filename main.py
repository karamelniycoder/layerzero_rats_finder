from openpyxl.styles import Color, PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.cell import get_column_letter
from openpyxl import Workbook, load_workbook
from aiohttp import ClientSession
from bs4 import BeautifulSoup
from datetime import datetime
from loguru import logger
from sys import stderr
from time import sleep
import asyncio
import os


logger.remove()
logger.add(stderr, format="<white>{time:HH:mm:ss}</white> | <level>{message}</level>")


class Excel:
    def __init__(self, total_len: int, name: str):
        if not os.path.isdir('results'): os.mkdir('results')

        self.file_name = f'{name}_{total_len}accs_{datetime.now().strftime("%d_%m_%Y_%H_%M_%S")}.xlsx'
        self.lock = asyncio.Lock()

        workbook = Workbook()
        sheet = workbook.active

        sheet['A2'] = 'Index'
        sheet['B2'] = 'Address'
        sheet['C2'] = 'Reports count'
        sheet['D2'] = 'Reports data'

        sheet.column_dimensions['A'].width = 10
        sheet.column_dimensions['B'].width = 46
        sheet.column_dimensions['C'].width = 24

        for cell in sheet._cells:
            sheet.cell(cell[0], cell[1]).font = Font(bold=True)
            sheet.cell(cell[0], cell[1]).alignment = Alignment(horizontal='center')
            sheet.cell(cell[0], cell[1]).border = Border(left=Side(style='thin'), bottom=Side(style='thin'), top=Side(style='thin'), right=Side(style='thin'))

        workbook.save('results/'+self.file_name)


    async def edit_table(self, index: str, address: str, address_result: dict):
        async with self.lock:
            while True:
                try:
                    workbook = load_workbook('results/' + self.file_name)
                    sheet = workbook.active

                    if address_result.get('status') == False:
                        valid_info = [
                            index,
                            address,
                            0
                        ]

                    elif address_result.get('status') == True:
                        valid_info = [
                            index,
                            address,
                            len(address_result["reports_data"]),
                            *[f'=HYPERLINK("{report["link"]}", "{report["text"]}")' for report in address_result["reports_data"]]
                        ]

                    sheet.append(valid_info)

                    for row_cells in sheet.iter_rows(min_row=sheet.max_row, max_row=sheet.max_row):
                        for cell in row_cells:
                            if cell.column == 3:
                                if int(cell.value) > 0: rgb_color = 'ff0f0f'
                                else: rgb_color = '32CD32'
                                cell.fill = PatternFill(patternType='solid', fgColor=Color(rgb=rgb_color))

                            elif cell.column == 4:
                                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'))

                            if cell.column > 3:
                                cell.font = Font(underline='single')

                    workbook.save('results/'+self.file_name)
                    return True
                except PermissionError:
                    logger.warning(f'Excel | Cant save excel file, close it!')
                    sleep(3)
                except Exception as err:
                    logger.critical(f'Excel | Cant save excel file: {err} | {address}')
                    return False


    def final_formatting(self):
        while True:
            try:
                ratted_addresses = 0
                workbook = load_workbook('results/' + self.file_name)
                sheet = workbook.active

                max_lengths = {}
                for row_cells in sheet.iter_rows(min_row=sheet.min_row, max_row=sheet.max_row, min_col=3):
                    for cell in row_cells:
                        if cell.column == 3:
                            if type(cell.value) == int and int(cell.value) > 0: ratted_addresses += 1
                            continue

                        if not max_lengths.get(get_column_letter(cell.column)): max_lengths[get_column_letter(cell.column)] = 0

                        if '=HYPERLINK' in str(cell.value): cell_len = len(str(cell.value).split(', "')[1])
                        else: cell_len = len(str(cell.value))

                        if cell_len > max_lengths[get_column_letter(cell.column)]: max_lengths[get_column_letter(cell.column)] = cell_len

                for column in max_lengths:
                    sheet.column_dimensions[column].width = max_lengths[column]

                sheet['C1'] = f'Total ratted addresses: {ratted_addresses}'
                sheet['C1'].font = Font(bold='single', italic='single')

                workbook.save('results/'+self.file_name)
                return True
            except PermissionError:
                logger.warning(f'Excel | Cant save excel file, close it!')
                sleep(3)
            except Exception as err:
                logger.critical(f'Excel | Cant save excel file: {err}')
                return False


async def find_address(address: str, sem: asyncio.Semaphore, excel: Excel, index: str):
    async with sem:
        async with ClientSession() as session:
            session.headers.update({
                'Referer': 'https://github.com/LayerZero-Labs/sybil-report/issues',
                'User-Agent': "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
            })

            while True:
                r = await session.get(f'https://github.com/LayerZero-Labs/sybil-report/issues?q={address}')
                response = await r.text()

                if 'No results matched your search.' in response:
                    logger.info(f'[+] {index} Address {address} is clear!')
                    result = {"status": False}
                    break

                elif 'You have exceeded a secondary rate limit' in response:
                    logger.warning(f'[-] Github Rate Limit, sleeping 60 seconds')
                    await asyncio.sleep(60)

                else:
                    soup = BeautifulSoup(response, "lxml")
                    reports_ = soup.findAll('div', class_="js-navigation-container js-active-navigation-container")

                    if not reports_:
                        logger.warning(f'[-] Something wrong with address {address}: {response}')
                        logger.info(f'[+] {index} Address {address} is clear!')
                        result = {"status": False}
                        break

                    reports = reports_[0].findAll('div', class_="Box-row Box-row--focus-gray p-0 mt-0 js-navigation-item js-issue-row")
                    reports_data = []
                    for report_ in reports:
                        report = report_.find('a', class_="Link--primary v-align-middle no-underline h4 js-navigation-open markdown-title")
                        reports_data.append({"link": "https://github.com" + report["href"], "text": report.text})

                    logger.error(f'[-] {index} Address {address} is ratted with {len(reports_data)} report(s)!')
                    result = {"status": True, "reports_data": reports_data}
                    break

            await excel.edit_table(index=index, address=address, address_result=result)


async def runner(addresses: list):
    logger.debug(f'[•] Starting parser for {len(addresses)} addresses...\n')
    excel = Excel(total_len=len(addresses), name="rats_parser")
    sem = asyncio.Semaphore(20)

    await asyncio.gather(*[find_address(address=address, sem=sem, excel=excel, index=f"[{index+1}/{len(addresses)}]") for index, address in enumerate(addresses)])

    excel.final_formatting()


if __name__ == "__main__":
    with open("addresses.txt") as f: addresses = f.read().splitlines()

    asyncio.run(runner(addresses=addresses))

    sleep(0.1)
    input('\nSee detailed info in folder `results`\n\n\t> Exit')
